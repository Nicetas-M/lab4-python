import csv
import openpyxl
from openpyxl import Workbook
from datetime import datetime

# Функція для обчислення віку
def calculate_age(birthdate):
    today = datetime.today()
    return today.year - birthdate.year - ((today.month, today.day) < (birthdate.month, birthdate.day))

# Створення нового Excel файлу
wb = Workbook()

# Створення аркуша "all"
ws_all = wb.active
ws_all.title = 'all'

# Додавання заголовків
ws_all.append(['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік'])

# Створення порожніх аркушів для вікових категорій
ws_younger_18 = wb.create_sheet('younger_18')
ws_18_45 = wb.create_sheet('18-45')
ws_45_70 = wb.create_sheet('45-70')
ws_older_70 = wb.create_sheet('older_70')

# Додавання заголовків до аркушів вікових категорій
for sheet in [ws_younger_18, ws_18_45, ws_45_70, ws_older_70]:
    sheet.append(['№', 'Прізвище', 'Ім’я', 'По батькові', 'Дата народження', 'Вік'])

# Читання даних з CSV файлу та категоризація за віком
with open('employees.csv', 'r', encoding='utf-8-sig') as file:
    reader = csv.reader(file)
    next(reader)  # Пропускаємо заголовок

    # Лічильники для кожної категорії
    idx_all = 1
    idx_younger_18 = 1
    idx_18_45 = 1
    idx_45_70 = 1
    idx_older_70 = 1

    for row in reader:
        birth_date = datetime.strptime(row[4], '%Y-%m-%d')  # Перетворюємо рядок на дату
        age = calculate_age(birth_date)  # Обчислюємо вік

        # Додавання співробітника до аркуша "all"
        ws_all.append([idx_all, row[0], row[1], row[2], row[4], age])
        idx_all += 1

        # Додавання співробітника до відповідного аркуша за віковою категорією
        if age < 18:
            ws_younger_18.append([idx_younger_18, row[0], row[1], row[2], row[4], age])
            idx_younger_18 += 1
        elif 18 <= age <= 45:
            ws_18_45.append([idx_18_45, row[0], row[1], row[2], row[4], age])
            idx_18_45 += 1
        elif 45 <= age <= 70:
            ws_45_70.append([idx_45_70, row[0], row[1], row[2], row[4], age])
            idx_45_70 += 1
        else:
            ws_older_70.append([idx_older_70, row[0], row[1], row[2], row[4], age])
            idx_older_70 += 1

# Збереження Excel файлу
try:
    wb.save('employees_categorized.xlsx')
    print("Ok, файл успішно створений")
except Exception as e:
    print(f"Помилка при створенні файлу: {e}")
